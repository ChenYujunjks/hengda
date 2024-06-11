import openpyxl
from helpers import second_star_index, get_scbh_from_shengchan

source_file = 'eg/input_source.xlsx'
shengchan_file = 'eg/help.xlsx'
output_file = 'include%/output.xlsx'
new_sheet_name = 'default_output'

def modify_excel_data(source_file, target_file, new_sheet_name):
    try:
        source_wb = openpyxl.load_workbook(source_file)
        source_sheet = source_wb.active  # Assume the data is in the first sheet
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return
    except Exception as e:
        print(f"An unexpected error occurred while loading the source file: {e}")
        return

    # Open the target file (create a new file if it doesn't exist)
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()

    # Create a new sheet in the target file
    if new_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)

    # Copy and modify data
    index_xh = 1
    for row_idx, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Handle header row
            new_row = ["序号", "生产编号", "开票日期", "发票号码", "客户名称", "货物名称", "备注软件名称", "数量", "不含税金额", "税额", "合计", "硬件成本"]
            target_sheet.append(new_row)
        else:
            # Data rows with filtering conditions
            try:
                SUM_IJ = float(row[16]) + float(row[18])
            except ValueError:
                SUM_IJ = None
                continue
            if row[26] is not None and isinstance(row[26], str) and "含" in row[26] and "%" in row[26]:  # Modify remark software name
                index1 = row[26].find("含")  # Find the "含" in the remark
                bzrjmc = row[26][index1:]  
                row6index = second_star_index(row[11])   # Second star index in product name
                hwmc = row[11][row6index+1:]

                # Get production number for the invoice number
                fphm = row[3]  # Assume invoice number is in the fourth column
                scbh = get_scbh_from_shengchan(shengchan_file, fphm)
                # Start writing new row
                new_row = [
                    index_xh, 
                    scbh if scbh else 0,  # If production number not found, fill 0,
                    row[8],  # Invoice date
                    fphm,  # Invoice number
                    row[7],   # Customer name
                    hwmc,  # Product name after second star index
                    bzrjmc, # Remark software name
                    row[14],  # Quantity
                    row[16], 
                    row[18], 
                    SUM_IJ
                ]
                index_xh += 1
                target_sheet.append(new_row)

    target_wb.save(target_file)

if __name__ == "__main__":
    modify_excel_data(source_file, output_file, new_sheet_name)