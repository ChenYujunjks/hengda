import openpyxl
from openpyxl.styles import PatternFill

# 打开两个Excel文件
wb_hand = openpyxl.load_workbook('hand.xlsx')
wb_target = openpyxl.load_workbook('target.xlsx')

# 选择工作表
sheet_hand = wb_hand.worksheets[0]  # 第一个工作表
sheet_target = wb_target.worksheets[1]  # 第二个工作表

# 获取数据
data_hand = {sheet_hand.cell(row=i, column=5).value: i for i in range(4, sheet_hand.max_row + 1)}
data_target = {sheet_target.cell(row=i, column=4).value: i for i in range(2, sheet_target.max_row + 1)}

# 创建一个新的工作簿和工作表
wb_output = openpyxl.Workbook()
sheet_output = wb_output.active

# 设置颜色填充
orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# 设置第一行标题，使用 target 表的标题
for col in range(1, sheet_target.max_column + 1):
    sheet_output.cell(row=1, column=col, value=sheet_target.cell(row=1, column=col).value)

# 写入表 target 中存在但表 hand 中不存在的行
output_row = 2  # 从第二行开始写入数据
for value_target, row_target in data_target.items():
    if value_target not in data_hand:
        for col in range(1, sheet_target.max_column + 1):
            cell = sheet_target.cell(row=row_target, column=col)
            new_cell = sheet_output.cell(row=output_row, column=col, value=cell.value)
            new_cell.fill = blue_fill
        output_row += 1

# 写入表 hand 中存在但表 target 中不存在的行，舍弃 hand 表的 C 列
for value_hand, row_hand in data_hand.items():
    if value_hand not in data_target:
        output_col = 1
        for col in range(1, sheet_hand.max_column + 1):
            if col != 3:  # 跳过 C 列
                cell = sheet_hand.cell(row=row_hand, column=col)
                new_cell = sheet_output.cell(row=output_row, column=output_col, value=cell.value)
                new_cell.fill = orange_fill
                output_col += 1
        output_row += 1

# 保存输出文件
wb_output.save('output3.xlsx')

print("Comparison completed. Results saved in 'output3.xlsx'.")
