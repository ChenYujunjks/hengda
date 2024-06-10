import openpyxl
from src.data_processing import save_excel  # 确保导入了 save_excel 函数

def second_star_index(text):
    first_star_index = text.find('*')
    if first_star_index == -1:
        return -1  # 如果没有找到第一个星号，返回-1
    
    second_star_index = text.find('*', first_star_index + 1)
    return second_star_index

def get_scbh_from_shengchan(shengchan_file, fphm):  # 获取生产编号
    shengchan_wb = openpyxl.load_workbook(shengchan_file)
    shengchan_sheet = shengchan_wb.active  # 假设生产文件中的数据在第一个工作表
    
    for row in shengchan_sheet.iter_rows(min_row=2, values_only=True):  # 假设第1行为标题行，从第2行开始
        if row[4] == fphm:  # E列是第五列（索引从0开始）
            return row[1]  # B列是第二列（索引从0开始）
    return None

def modify_excel_data(source_file, shengchan_file, target_file=None):
    if target_file is None:
        target_file = 'output/output.xlsx'
    
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active  # 假设源文件中需要复制的数据在第一个工作表

    # 打开目标文件（如果不存在则创建一个新的文件）
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()

    new_sheet_name = 'ModifiedSheet'
    if new_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)

    # 复制并修改数据
    index_xh = 1
    for row_idx, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            new_row = ["序号", "生产编号", "开票日期", "发票号码", "客户名称", "货物名称", "备注软件名称", "数量", "不含税金额", "税额", "合计", "硬件成本"]
            target_sheet.append(new_row)
        else:
            try:
                SUM_IJ = float(row[16]) + float(row[18])
            except ValueError:
                SUM_IJ = None
                continue
            if "含" in row[26]:
                index1 = row[26].find("含")
                bzrjmc = row[26][index1:]
                row6index = second_star_index(row[11])
                hwmc = row[11][row6index+1:]

                fphm = row[3]
                scbh = get_scbh_from_shengchan(shengchan_file, fphm)
                new_row = [
                    index_xh,
                    scbh if scbh else 0,
                    row[8],
                    fphm,
                    row[7],
                    hwmc,
                    bzrjmc,
                    row[14],
                    row[16],
                    row[18],
                    SUM_IJ
                ]
                index_xh += 1
                target_sheet.append(new_row)

    save_excel(target_wb, target_file)
