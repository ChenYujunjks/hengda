import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict

def compare_excel_files(target_file, hand_file, output_file):
    wb_hand = openpyxl.load_workbook(hand_file)
    wb_target = openpyxl.load_workbook(target_file)

    sheet_hand = wb_hand.worksheets[0]
    sheet_target = wb_target.worksheets[1]

    wb_output = openpyxl.Workbook()
    sheet_output = wb_output.active

    sheet_output.cell(row=1, column=1, value="发票号码")
    sheet_output.cell(row=1, column=2, value="Target 不含税金额总和")
    sheet_output.cell(row=1, column=3, value="Hand 不含税金额总和")
    sheet_output.cell(row=1, column=4, value="差值 (Target - Hand)")

    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    hand_invoices = defaultdict(float)
    for i in range(4, sheet_hand.max_row + 1):
        invoice_number = sheet_hand.cell(row=i, column=5).value
        amount = sheet_hand.cell(row=i, column=10).value
        if isinstance(amount, (int, float)):
            hand_invoices[invoice_number] += amount

    target_invoices = defaultdict(float)
    for i in range(2, sheet_target.max_row + 1):
        invoice_number = sheet_target.cell(row=i, column=4).value
        amount = sheet_target.cell(row=i, column=9).value
        if isinstance(amount, (int, float)):
            target_invoices[invoice_number] += amount

    positive_diff_rows = []
    zero_diff_rows = []
    negative_diff_rows = []

    for invoice_number in hand_invoices.keys() | target_invoices.keys():
        hand_total = hand_invoices[invoice_number]
        target_total = target_invoices[invoice_number]
        difference = target_total - hand_total

        if abs(difference) < 1:
            difference = 0

        row_data = (invoice_number, target_total, hand_total, difference)

        if difference > 0:
            positive_diff_rows.append(row_data)
        elif difference < 0:
            negative_diff_rows.append(row_data)
        else:
            zero_diff_rows.append(row_data)

    output_row = 2
    for row_data in positive_diff_rows:
        for col, value in enumerate(row_data, start=1):
            new_cell = sheet_output.cell(row=output_row, column=col, value=value)
            if col == 4:
                new_cell.fill = blue_fill
        output_row += 1

    for row_data in negative_diff_rows:
        for col, value in enumerate(row_data, start=1):
            new_cell = sheet_output.cell(row=output_row, column=col, value=value)
            if col == 4:
                new_cell.fill = orange_fill
        output_row += 1

    for row_data in zero_diff_rows:
        for col, value in enumerate(row_data, start=1):
            sheet_output.cell(row=output_row, column=col, value=value)
        output_row += 1

    wb_output.save(output_file)
