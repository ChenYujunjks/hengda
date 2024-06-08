import os
import openpyxl
from data_processing import process_data

# 构建相对路径
source_file = os.path.join('..', 'test', 'source.xlsx')
target_file = os.path.join('..', 'test', 'target.xlsx')
new_sheet_name = 'try251'

def get_scbh_from_shengchan(shengchan_file, fphm): #获取生产编号
    print(f"Reading {shengchan_file} to find SCBH for FPHM: {fphm}")
    shengchan_wb = openpyxl.load_workbook(shengchan_file)
    shengchan_sheet = shengchan_wb.active  # 假设生产文件中的数据在第一个工作表
    
    for row in shengchan_sheet.iter_rows(min_row=2, values_only=True):  # 假设第1行为标题行，从第2行开始
        if row[4] == fphm:  # E列是第五列（索引从0开始）
            print(f"Found SCBH: {row[1]} for FPHM: {fphm}")
            return row[1]  # B列是第二列（索引从0开始）
    return None

def modify_excel_data(source_file, target_file, new_sheet_name):
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active  # 假设源文件中需要复制的数据在第一个工作表

    # 打开目标文件（如果不存在则创建一个新的文件）
    try:
        target_wb = openpyxl.load_workbook(target_file)
        print("Target workbook loaded.")
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()
        print("Target workbook not found. New workbook created...")

# 在目标文件中创建一个新的工作表
    if new_sheet_name in target_wb.sheetnames:
        print(f"Sheet '{new_sheet_name}' already exists in target workbook.")
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)
        print(f"Sheet '{new_sheet_name}' created.")
        
    process_data(source_sheet, target_sheet)
    target_wb.save(target_file)
    

if __name__ == "__main__":
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' not found.")
    else:
        print("Source workbook loaded.")
        modify_excel_data(source_file, target_file, new_sheet_name)
