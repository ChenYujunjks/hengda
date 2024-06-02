import openpyxl

source_file = 'source.xlsx'
target_file = 'target.xlsx'
new_sheet_name = 'try2'

def copy_and_modify_excel_data(source_file, target_file, new_sheet_name):
    
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active  # 假设源文件中需要复制的数据在第一个工作表

    # 打开目标文件（如果不存在则创建一个新的文件）
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()

# 在目标文件中创建一个新的工作表
    if new_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)

# 复制并修改数据
    for row_idx, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # 处理标题行
            new_row = ["序号", "生产编号", row[5], row[2], "客户名称", "货物名称", "备注软件名称","数量", "不含税金额","税额", "合计"]
        else:
            SUM_IJ = row[8] + row[10] if isinstance(row[9], (int, float)) and isinstance(row[11], (int, float)) else None
            # 处理数据行
            new_row = [
                row_idx - 1, 
                0,
                row[5], 
                row[2], 
                row[4], 
                row[6], 
                row[17], 
                1,
                row[9], 
                row[11], 
                SUM_IJ
            ]
        target_sheet.append(new_row)

    # 保存目标文件
    target_wb.save(target_file)

copy_and_modify_excel_data(source_file, target_file, new_sheet_name)
