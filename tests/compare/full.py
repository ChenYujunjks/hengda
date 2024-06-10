import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict

# 打开两个Excel文件
wb_hand = openpyxl.load_workbook('../hand.xlsx')
wb_target = openpyxl.load_workbook('../target.xlsx')

# 选择工作表
sheet_hand = wb_hand.worksheets[0]  # 第一个工作表
sheet_target = wb_target.worksheets[1]  # 第二个工作表

# 创建一个新的工作簿和工作表
wb_output = openpyxl.Workbook()
sheet_output = wb_output.active

# 设置第一行标题
sheet_output.cell(row=1, column=1, value="发票号码")
sheet_output.cell(row=1, column=2, value="Target 不含税金额总和")
sheet_output.cell(row=1, column=3, value="Hand 不含税金额总和")
sheet_output.cell(row=1, column=4, value="差值 (Target - Hand)")

# 设置颜色填充
orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# 获取hand表的发票号码和不含税金额
hand_invoices = defaultdict(float)
for i in range(4, sheet_hand.max_row + 1):
    invoice_number = sheet_hand.cell(row=i, column=5).value
    amount = sheet_hand.cell(row=i, column=10).value  # J列是第10列
    if isinstance(amount, (int, float)):
        hand_invoices[invoice_number] += amount

# 获取target表的发票号码和不含税金额
target_invoices = defaultdict(float)
for i in range(2, sheet_target.max_row + 1):
    invoice_number = sheet_target.cell(row=i, column=4).value
    amount = sheet_target.cell(row=i, column=9).value  # I列是第9列
    if isinstance(amount, (int, float)):
        target_invoices[invoice_number] += amount

# 分别存储差值大于0、等于0、小于0的行
positive_diff_rows = []
zero_diff_rows = []
negative_diff_rows = []

# 比较发票号码并分类存储结果
for invoice_number in hand_invoices.keys() | target_invoices.keys():
    hand_total = hand_invoices[invoice_number]
    target_total = target_invoices[invoice_number]
    difference = target_total - hand_total
    
    # 视差值小于1的为0
    if abs(difference) < 1:
        difference = 0
    
    row_data = (invoice_number, target_total, hand_total, difference)
    
    if difference > 0:
        positive_diff_rows.append(row_data)
    elif difference < 0:
        negative_diff_rows.append(row_data)
    else:
        zero_diff_rows.append(row_data)

# 写入差值大于0的行，背景色为蓝色
output_row = 2
for row_data in positive_diff_rows:
    for col, value in enumerate(row_data, start=1):
        new_cell = sheet_output.cell(row=output_row, column=col, value=value)
        if col == 4:
            new_cell.fill = blue_fill
    output_row += 1

# 写入差值小于0的行，背景色为橘色
for row_data in negative_diff_rows:
    for col, value in enumerate(row_data, start=1):
        new_cell = sheet_output.cell(row=output_row, column=col, value=value)
        if col == 4:
            new_cell.fill = orange_fill
    output_row += 1

# 写入差值等于0的行，不设置背景色
for row_data in zero_diff_rows:
    for col, value in enumerate(row_data, start=1):
        sheet_output.cell(row=output_row, column=col, value=value)
    output_row += 1

# 保存输出文件
wb_output.save('output_comparison.xlsx')

print("Comparison completed. Results saved in 'output_comparison.xlsx'.")