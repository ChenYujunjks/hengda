from openpyxl import load_workbook, Workbook
import os

# 指定文件路径
source_file = "/Users/bruce/code2024summer/python-react/openecel/target.xlsx"
new_file = "/Users/bruce/code2024summer/python-react/openecel/source_first_70_rows.xlsx"

# 检查文件是否存在
if not os.path.exists(source_file):
    print(f"文件 {source_file} 不存在，请检查文件路径")
else:
    try:
        # 加载现有的 Excel 文件
        wb = load_workbook(source_file)
        ws = wb.active

        # 创建一个新的工作簿
        new_wb = Workbook()
        new_ws = new_wb.active

        # 复制前 70 行的数据到新的工作簿
        for row in ws.iter_rows(min_row=1, max_row=70):
            for cell in row:
                new_ws[cell.coordinate] = cell.value

        # 保存新的 Excel 文件
        new_wb.save(new_file)
        print(f"前 70 行的数据已保存到 {new_file}")
    except Exception as e:
        print(f"处理文件时出现错误: {e}")
