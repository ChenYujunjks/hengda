import openpyxl

# 使用示例
source_file = 'source.xlsx'
target_file = 'new_target.xlsx'
new_sheet_name = 'Sheet2'

def copy_excel_data(source_file, target_file, new_sheet_name):
    try:
        source_wb = openpyxl.load_workbook(source_file)
    except Exception as e:
        print(f"Error loading source file: {e}")
        return
    
    source_sheet = source_wb.active  # 假设源文件中需要复制的数据在第一个工作表

    # 打开目标文件（如果不存在则创建一个新的文件）
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()
    except Exception as e:
        print(f"Error loading target file: {e}")
        return

    # 在目标文件中创建一个新的工作表
    if new_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)

    # 复制所有行的数据
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # 保存目标文件
    try:
        target_wb.save(target_file)
        print(f"数据已成功复制到 {target_file} 的 {new_sheet_name} 工作表中")
    except Exception as e:
        print(f"Error saving target file: {e}")


copy_excel_data(source_file, target_file, new_sheet_name)
