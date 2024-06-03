import openpyxl

source_file = 'input_source.xlsx'
output_file = 'source.xlsx'
max_rows = 600

def export_top_rows(source_file, output_file, max_rows):
    # 打开源文件
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source.active

    # 创建一个新的工作簿和工作表
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "TopRows"

    # 复制前max_rows行数据
    for row in ws_source.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        ws_output.append(row)

    # 保存到新的文件
    wb_output.save(output_file)
    print(f"Exported top {max_rows} rows to {output_file}")

export_top_rows(source_file, output_file, max_rows)
