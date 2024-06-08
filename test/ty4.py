import openpyxl

source_file = 'source.xlsx'
shengchan_file = 'help.xlsx'
target_file = 'target.xlsx'
new_sheet_name = 'try4'

def second_star_index(text):
    first_star_index = text.find('*')
    if first_star_index == -1:
        return -1  # 如果没有找到第一个星号，返回-1
    
    second_star_index = text.find('*', first_star_index + 1)
    return second_star_index

def get_scbh_from_shengchan(shengchan_file, fphm): #获取生产编号
    print(f"Reading {shengchan_file} to find SCBH for FPHM: {fphm}")
    shengchan_wb = openpyxl.load_workbook(shengchan_file)
    shengchan_sheet = shengchan_wb.active  # 假设生产文件中的数据在第一个工作表
    
    for row in shengchan_sheet.iter_rows(min_row=2, values_only=True):  # 假设第1行为标题行，从第2行开始
        if row[4] == fphm:  # E列是第五列（索引从0开始）
            print(f"Found SCBH: {row[1]} for FPHM: {fphm}")
            return row[1]  # B列是第二列（索引从0开始）
    return None

def modify_excel_data(source_file, target_file, new_sheet_name):
    
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active  # 假设源文件中需要复制的数据在第一个工作表

    # 打开目标文件（如果不存在则创建一个新的文件）
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()

# 在目标文件中创建一个新的工作表
    if new_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[new_sheet_name]
    else:
        target_sheet = target_wb.create_sheet(title=new_sheet_name)

# 复制并修改数据
    index_xh = 1
    for row_idx, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # 处理标题行
            new_row = ["序号", "生产编号", '开票日期', "发票号码", "客户名称", "货物名称", "备注软件名称", "数量", "不含税金额","税额", "合计"]
            target_sheet.append(new_row)
        else:
            # 数据行 条件筛选
            try:
                SUM_IJ = float(row[16]) + float(row[18])
            except ValueError:
                SUM_IJ = None
                continue
            if "含" in row[26]:   #改动 备注软件名称
                index1 = row[26].find("含")  # 找到转账备注中的 含：恒达富士乘客电梯变频控制软件V1.0的数据
                bzrjmc = row[26][index1:]  
                row6index = second_star_index(row[11])   # 货物名称 中第二个星号的index
                hwmc = row[11][row6index+1:]

                 # 获取发票号码对应的生产编号
                fphm = row[3]  # 假设发票号码在第四列
                scbh = get_scbh_from_shengchan(shengchan_file, fphm)
            # 开始写 new row
                new_row = [
                    index_xh, 
                    scbh if scbh else 0,  # 如果找不到生产编号，则填0,
                    row[8],  # 开票日期
                    fphm,  # 发票号码
                    row[7],   # 客户名称
                    hwmc,  # 货物名称 中第二个星号的index
                    bzrjmc, # 备注软件名称
                    row[14],  # 数量
                    row[16], 
                    row[18], 
                    SUM_IJ
                ]
                index_xh += 1
                target_sheet.append(new_row)

    target_wb.save(target_file)

modify_excel_data(source_file, target_file, new_sheet_name)